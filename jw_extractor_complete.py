import requests
from bs4 import BeautifulSoup
import re
import json
from typing import Optional, Dict, List, Tuple
from datetime import datetime
import pandas as pd

# Para Google Sheets
try:
    from google.colab import auth
    import gspread
    from google.auth import default
    SHEETS_DISPONIBLE = True
except Exception as e:
    print(f"Error importing gspread or google.colab: {e}")
    SHEETS_DISPONIBLE = False

# Para subir/descargar archivos en Colab
try:
    from google.colab import files
    IN_COLAB = True
except ImportError:
    IN_COLAB = False

# ==================== CONFIGURACIÓN ====================
HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    'Accept-Language': 'es-ES,es;q=0.9',
    'Connection': 'keep-alive'
}

TIMEOUT = 30
MAX_REINTENTOS = 3

LIBROS_BIBLIA = (
    'ECLESIASTÉS', 'GÉNESIS', 'ÉXODO', 'LEVÍTICO', 'NÚMEROS', 'DEUTERONOMIO',
    'JOSUÉ', 'JUECES', 'RUT', 'SAMUEL', 'REYES', 'CRÓNICAS', 'ESDRAS',
    'NEHEMÍAS', 'ESTER', 'JOB', 'SALMOS', 'PROVERBIOS', 'CANTARES',
    'ISAÍAS', 'JEREMÍAS', 'LAMENTACIONES', 'EZEQUIEL', 'DANIEL',
    'OSEAS', 'JOEL', 'AMÓS', 'ABDÍAS', 'JONÁS', 'MIQUEAS', 'NAHÚM',
    'HABACUC', 'SOFONÍAS', 'HAGEO', 'ZACARÍAS', 'MALAQUÍAS',
    'MATEO', 'MARCOS', 'LUCAS', 'JUAN', 'HECHOS', 'ROMANOS',
    'CORINTIOS', 'GÁLATAS', 'EFESIOS', 'FILIPENSES', 'COLOSENSES',
    'TESALONICENSES', 'TIMOTEO', 'TITO', 'FILEMÓN', 'HEBREOS',
    'SANTIAGO', 'PEDRO', 'JUDAS', 'APOCALIPSIS'
)

PATRONES = {
    'fecha': re.compile(r'\d{1,2}\s*(?:-\s*\d{1,2}|de\s+\w+\s+(?:a|al)\s+\d{1,2})\s+de\s+\w+', re.IGNORECASE),
    'cancion': re.compile(r'Canción\s+(\d+)', re.IGNORECASE),
    'palabras': re.compile(r'Palabras\s+de\s+(introducción|conclusión)\s*[:\(]?\s*(\d+)\s*min', re.IGNORECASE),
    'parte_numerada': re.compile(r'^(\d+)\.\s*([^\n(]+?)\s*\((\d+)\s*min', re.MULTILINE | re.IGNORECASE),
    'parte_sin_numero': re.compile(r'^(Empiece conversaciones|Haga revisitas|Estudio bíblico|Necesidades de la congregación|Canción del Reino y oración final)\s*\(?\s*(\d+)\s*min', re.MULTILINE | re.IGNORECASE)
}

# ==================== FUNCIONES PARA EXTRAER ENLACES ====================
def obtener_enlaces_semanas(url_indice: str) -> List[Dict[str, str]]:
    """Extrae todos los enlaces de semanas desde la URL índice."""
    try:
        print("🔍 Buscando todas las semanas disponibles...\n")
        response = requests.get(url_indice, headers=HEADERS, timeout=TIMEOUT)
        response.raise_for_status()
        soup = BeautifulSoup(response.content, 'html.parser')
        
        enlaces = []
        
        main_content_div = soup.find('div', class_='docPart')
        if main_content_div:
            links = main_content_div.find_all('a', href=True)
        else:
            links = soup.find_all('a', href=True)
        
        for link in links:
            href = link.get('href')
            texto = link.get_text(strip=True)
            
            if '/es/biblioteca/guia-actividades-reunion-testigos-jehova/' in href and texto:
                if href != url_indice and not href.endswith('/mwb/'):
                    if PATRONES['fecha'].search(texto):
                        if not href.startswith('http'):
                            href = f"https://www.jw.org{href}"
                        enlaces.append({'titulo': texto, 'url': href})
        
        enlaces.sort(key=lambda x: extraer_fecha_para_ordenar(x['titulo']))
        
        print(f"✅ Se encontraron {len(enlaces)} semanas\n")
        return enlaces
        
    except Exception as e:
        print(f"❌ Error al obtener enlaces: {e}")
        return []

def extraer_fecha_para_ordenar(titulo: str) -> tuple:
    """Extrae la fecha inicial para ordenar cronológicamente."""
    meses = {
        'enero': 1, 'febrero': 2, 'marzo': 3, 'abril': 4,
        'mayo': 5, 'junio': 6, 'julio': 7, 'agosto': 8,
        'septiembre': 9, 'octubre': 10, 'noviembre': 11, 'diciembre': 12
    }
    
    match = re.search(r'(\d{1,2})[- ].*?de\s+(\w+)', titulo, re.IGNORECASE)
    if match:
        dia = int(match.group(1))
        mes_texto = match.group(2).lower()
        mes = meses.get(mes_texto, 0)
        return (mes, dia)
    return (0, 0)

def mostrar_semanas_disponibles(enlaces: List[Dict[str, str]]) -> None:
    """Muestra las semanas disponibles."""
    if not enlaces:
        print("❌ No se encontraron semanas")
        return
    
    print("📅 SEMANAS DISPONIBLES:\n")
    for i, sem in enumerate(enlaces, 1):
        print(f"  {i}. {sem['titulo']}")
    print()

# ==================== FUNCIONES DE EXTRACCIÓN ====================
def obtener_contenido(url: str) -> Optional[str]:
    """Descarga y extrae texto de la página web con reintentos."""
    for intento in range(1, MAX_REINTENTOS + 1):
        try:
            response = requests.get(
                url,
                headers=HEADERS,
                timeout=TIMEOUT,
                allow_redirects=True
            )
            response.raise_for_status()
            soup = BeautifulSoup(response.content, 'html.parser')
            main = soup.find('main') or soup
            return main.get_text(separator='\n', strip=True)
        except requests.Timeout:
            if intento == MAX_REINTENTOS:
                print(f"⏱️ Timeout")
        except requests.RequestException as e:
            if intento == MAX_REINTENTOS:
                print(f"❌ Error: {e}")
    return None

def buscar_patron(contenido: str, patron: re.Pattern) -> str:
    """Busca un patrón y retorna el primer match completo."""
    match = patron.search(contenido)
    return match.group(0) if match else ''

def extraer_fecha_correcta(contenido: str) -> str:
    """Extrae la fecha correcta del contenido."""
    lineas = contenido.split('\n')
    
    for linea in lineas[:20]:
        fecha_match = PATRONES['fecha'].search(linea)
        if fecha_match:
            return fecha_match.group(0).strip()
    
    fecha_match = PATRONES['fecha'].search(contenido)
    if fecha_match:
        return fecha_match.group(0).strip()
    
    return ''

def extraer_lectura_biblica(contenido: str) -> str:
    """Extrae la lectura bíblica del contenido."""
    for libro in LIBROS_BIBLIA:
        patron = rf'({libro})\s*\d+(?::\d+)?(?:[-–]\d+(?::\d+)?)?'
        match = re.search(patron, contenido, re.IGNORECASE | re.DOTALL)
        if match:
            return re.sub(r'\s+', ' ', match.group(0)).strip()
    
    match2 = re.search(
        r'Lectura\s+b[ií]blica\s*[:\-]?\s*([A-Za-zÁÉÍÓÚáéíóúñÑ0-9\s:–\-]+)',
        contenido, re.IGNORECASE
    )
    return re.sub(r'\s+', ' ', match2.group(1)).strip() if match2 else ''

def extraer_canciones(contenido: str) -> Dict[str, str]:
    """Extrae números de las 3 canciones."""
    nums = PATRONES['cancion'].findall(contenido)
    return {
        'cancion_inicial': f"Canción {nums[0]}" if len(nums) > 0 else '',
        'cancion_intermedia': f"Canción {nums[1]}" if len(nums) > 1 else '',
        'cancion_final': f"Canción {nums[2]}" if len(nums) > 2 else ''
    }

def extraer_palabras(contenido: str) -> Dict[str, str]:
    """Extrae palabras de introducción y conclusión."""
    palabras = {}
    for match in PATRONES['palabras'].finditer(contenido):
        tipo, mins = match.groups()
        palabras[f'palabras_{tipo}'] = f"Palabras de {tipo} ({mins} min)"
    
    return {
        'palabras_introduccion': palabras.get('palabras_introducción', ''),
        'palabras_conclusion': palabras.get('palabras_conclusión', '')
    }

def encontrar_posicion_cancion_intermedia(contenido: str) -> int:
    """Encuentra después de qué número de parte viene la canción intermedia."""
    canciones = list(PATRONES['cancion'].finditer(contenido))
    if len(canciones) < 2:
        return 6
    
    pos_cancion = canciones[1].start()
    partes_antes = []
    
    for match in PATRONES['parte_numerada'].finditer(contenido):
        if match.start() < pos_cancion:
            partes_antes.append(int(match.group(1)))
    
    return max(partes_antes) if partes_antes else 6

def extraer_partes(contenido: str) -> Tuple[Dict[str, List[Dict]], int]:
    """Extrae y clasifica partes dinámicamente según posición de canciones."""
    parte_antes_cancion = encontrar_posicion_cancion_intermedia(contenido)
    
    secciones = {
        'tesoros_biblia': [],
        'seamos_maestros': [],
        'vida_cristiana': []
    }
    
    contador_parte = 0
    
    # Extraer partes numeradas
    for match in PATRONES['parte_numerada'].finditer(contenido):
        num = int(match.group(1))
        contador_parte += 1
        parte = {
            'numero': contador_parte,
            'titulo': match.group(2).strip(),
            'duracion': f"{match.group(3)} min"
        }
        
        if num <= 3:
            secciones['tesoros_biblia'].append(parte)
        elif num <= parte_antes_cancion:
            secciones['seamos_maestros'].append(parte)
        else:
            secciones['vida_cristiana'].append(parte)
    
    # Extraer partes sin número (típicamente en "Vida cristiana")
    for match in PATRONES['parte_sin_numero'].finditer(contenido):
        contador_parte += 1
        parte = {
            'numero': contador_parte,
            'titulo': match.group(1).strip(),
            'duracion': f"{match.group(2)} min"
        }
        secciones['vida_cristiana'].append(parte)
    
    return secciones, parte_antes_cancion

def extraer_datos_reunion(url: str) -> Optional[Dict]:
    """Extrae todos los datos de la reunión desde la URL."""
    contenido = obtener_contenido(url)
    if not contenido:
        return None
    
    partes_data, corte = extraer_partes(contenido)
    
    datos = {
        'fecha': extraer_fecha_correcta(contenido),
        'lectura_biblica': extraer_lectura_biblica(contenido),
        **extraer_canciones(contenido),
        **extraer_palabras(contenido),
        **partes_data,
        '_corte_cancion': corte
    }
    
    # DEBUG: Mostrar qué se extrajo
    print(f"  📅 Fecha: {datos['fecha']}")
    print(f"  📋 Partes encontradas: Tesoros={len(datos['tesoros_biblia'])}, Maestros={len(datos['seamos_maestros'])}, Vida={len(datos['vida_cristiana'])}")
    print()
    
    return datos

# ==================== FUNCIONES PARA GOOGLE SHEETS ====================
def conectar_google_sheets():
    """Conecta con Google Sheets y retorna el cliente."""
    if not SHEETS_DISPONIBLE:
        print("❌ gspread no está disponible. Instala con: !pip install gspread")
        return None
    
    try:
        print("🔐 Autenticando con Google...\n")
        auth.authenticate_user()
        creds, _ = default()
        gc = gspread.authorize(creds)
        print("✅ Autenticación exitosa\n")
        return gc
    except Exception as e:
        print(f"❌ Error de autenticación: {e}")
        return None

def crear_plantilla_sheets(gc, titulo_libro: str) -> Optional[str]:
    """Crea una nueva hoja de cálculo con la plantilla."""
    try:
        print(f"📝 Creando plantilla: '{titulo_libro}'...\n")
        
        # Crear libro
        sh = gc.create(titulo_libro)
        worksheet = sh.sheet1
        worksheet.update_title("Reuniones")
        
        # Encabezados expandidos para 9 partes
        encabezados = [
            'Semana', 'Fecha', 'Lectura Bíblica',
            'Canción Inicial', 'Palabras Introducción',
        ]
        
        # Agregar 9 pares de columnas para partes
        for i in range(1, 10):
            encabezados.extend([f'Parte {i}', f'Duración {i}'])
        
        encabezados.extend([
            'Canción Intermedia', 'Palabras Conclusión', 'Canción Final'
        ])
        
        worksheet.append_row(encabezados)
        
        # Formatear encabezado
        worksheet.format('1:1', {
            'backgroundColor': {'red': 0.2, 'green': 0.2, 'blue': 0.6},
            'textFormat': {'foregroundColor': {'red': 1, 'green': 1, 'blue': 1}, 'bold': True}
        })
        
        # Fijar encabezado
        worksheet.freeze(rows=1)
        
        # Compartir (opcional)
        sh.share('', perm_type='anyone', role='reader')
        
        print(f"✅ Plantilla creada: {sh.url}\n")
        return sh.id
        
    except Exception as e:
        print(f"❌ Error al crear plantilla: {e}")
        return None

def rellenar_sheets(gc, spreadsheet_id: str, datos_lista: List[Dict]) -> None:
    """Rellena la plantilla de Sheets con los datos."""
    try:
        print("📥 Rellenando Google Sheets...\n")
        sh = gc.open_by_key(spreadsheet_id)
        worksheet = sh.sheet1
        
        for i, datos in enumerate(datos_lista, 1):
            partes = [
                *datos['tesoros_biblia'],
                *datos['seamos_maestros'],
                *datos['vida_cristiana']
            ]
            partes.sort(key=lambda x: x['numero'])
            
            # Rellenar fila
            fila = [
                i,
                datos['fecha'],
                datos['lectura_biblica'],
                datos['cancion_inicial'],
                datos['palabras_introduccion'],
            ]
            
            # Agregar partes (máximo 9)
            for j in range(9):
                if j < len(partes):
                    fila.append(partes[j]['titulo'])
                    fila.append(partes[j]['duracion'])
                else:
                    fila.append('')
                    fila.append('')
            
            fila.extend([
                datos['cancion_intermedia'],
                datos['palabras_conclusion'],
                datos['cancion_final']
            ])
            
            worksheet.append_row(fila)
        
        print(f"✅ {len(datos_lista)} semanas añadidas a Sheets\n")
        print(f"📊 Accede aquí: {sh.url}\n")
        
    except Exception as e:
        print(f"❌ Error al rellenar Sheets: {e}")
        return None

def crear_plantilla_excel_local(nombre: str = "plantilla_reuniones.xlsx") -> None:
    """Crea una plantilla Excel vacía para descargar."""
    try:
        print(f"📝 Creando plantilla Excel: '{nombre}'...\n")
        
        encabezados = [
            'Semana', 'Fecha', 'Lectura Bíblica',
            'Canción Inicial', 'Palabras Introducción',
        ]
        
        # Agregar 9 pares de columnas para partes
        for i in range(1, 10):
            encabezados.extend([f'Parte {i}', f'Duración {i}'])
        
        encabezados.extend([
            'Canción Intermedia', 'Palabras Conclusión', 'Canción Final'
        ])
        
        # Crear DataFrame vacío
        df = pd.DataFrame(columns=encabezados)
        
        # Guardar en Excel con formato
        with pd.ExcelWriter(nombre, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Reuniones', index=False)
            
            # Formatear
            from openpyxl.styles import PatternFill, Font, Alignment
            workbook = writer.book
            worksheet = writer.sheets['Reuniones']
            
            # Encabezado azul
            fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
            font = Font(bold=True, color='FFFFFF')
            
            for cell in worksheet[1]:
                cell.fill = fill
                cell.font = font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Ajustar ancho
            worksheet.column_dimensions['A'].width = 10
            worksheet.column_dimensions['B'].width = 18
            worksheet.column_dimensions['C'].width = 22
        
        print(f"✅ Plantilla Excel creada: {nombre}\n")
        
        if IN_COLAB:
            files.download(nombre)
            
    except Exception as e:
        print(f"❌ Error al crear Excel: {e}")

# ==================== FUNCIÓN PRINCIPAL ====================
def main():
    """Función principal con opción de crear plantilla."""
    print("\n" + "="*70)
    print("🚀 EXTRACTOR DE REUNIONES JW.ORG")
    print("="*70)
    print()
    
    # Opción 1: Solo descargar plantilla Excel
    print("¿Qué deseas hacer?")
    print("1. Descargar plantilla Excel vacía")
    print("2. Extraer datos a Google Sheets (requiere autenticación)")
    print("3. Extraer datos a Excel local")
    print()
    
    if IN_COLAB:
        opcion = input("Selecciona opción (1-3): ").strip()
    else:
        opcion = "3"
    
    if opcion == "1":
        crear_plantilla_excel_local("plantilla_reuniones.xlsx")
        print("💡 Descarga la plantilla, llénala manualmente y luego:")
        print("   1. Sube el archivo a Colab")
        print("   2. Importa los datos")
        return
    
    # Obtener datos
    URL_INDICE = input("Ingresar URL: ")
    
    enlaces = obtener_enlaces_semanas(URL_INDICE)
    
    if not enlaces:
        print("❌ No se encontraron semanas")
        return
    
    mostrar_semanas_disponibles(enlaces)
    
    # Procesar todas las semanas
    datos_todas = []
    errores = []
    
    for i, semana in enumerate(enlaces, 1):
        try:
            print(f"⏳ [{i}/{len(enlaces)}] {semana['titulo']}...", end=" ")
            datos = extraer_datos_reunion(semana['url'])
            
            if datos:
                print("✅")
                datos_todas.append(datos)
            else:
                print("❌")
                errores.append(semana['titulo'])
                
        except Exception as e:
            print(f"❌ ({e})")
            errores.append(f"{semana['titulo']}: {e}")
    
    print()
    print("="*70)
    print(f"✅ PROCESADAS: {len(datos_todas)}/{len(enlaces)}")
    if errores:
        print(f"❌ ERRORES: {len(errores)}")
    print("="*70)
    print()
    
    if not datos_todas:
        print("❌ No hay datos para guardar")
        return
    
    # Guardar según opción
    if opcion == "2" and SHEETS_DISPONIBLE:
        gc = conectar_google_sheets()
        if gc:
            titulo = input("\n¿Nombre para la hoja de cálculo? (default: Reuniones JW): ").strip()
            if not titulo:
                titulo = "Reuniones JW"
            
            spreadsheet_id = crear_plantilla_sheets(gc, titulo)
            if spreadsheet_id:
                rellenar_sheets(gc, spreadsheet_id, datos_todas)
    
    else:
        # Opción 3: Excel local
        nombre_archivo = "reuniones_datos.xlsx"
        print(f"💾 Generando {nombre_archivo}...\n")
        
        with pd.ExcelWriter(nombre_archivo, engine='openpyxl') as writer:
            filas = []
            
            for i, datos in enumerate(datos_todas, 1):
                partes = [*datos['tesoros_biblia'], *datos['seamos_maestros'], *datos['vida_cristiana']]
                partes.sort(key=lambda x: x['numero'])
                
                fila = [i, datos['fecha'], datos['lectura_biblica'], datos['cancion_inicial'], datos['palabras_introduccion']]
                
                # Agregar partes (máximo 9)
                for j in range(9):
                    if j < len(partes):
                        fila.append(partes[j]['titulo'])
                        fila.append(partes[j]['duracion'])
                    else:
                        fila.append('')
                        fila.append('')
                
                fila.extend([datos['cancion_intermedia'], datos['palabras_conclusion'], datos['cancion_final']])
                filas.append(fila)
            
            encabezados = ['Semana', 'Fecha', 'Lectura Bíblica', 'Canción Inicial', 'Palabras Introducción']
            
            # Agregar 9 pares de columnas para partes
            for i in range(1, 10):
                encabezados.extend([f'Parte {i}', f'Duración {i}'])
            
            encabezados.extend(['Canción Intermedia', 'Palabras Conclusión', 'Canción Final'])
            
            df = pd.DataFrame(filas, columns=encabezados)
            df.to_excel(writer, sheet_name='Reuniones', index=False)
            
            # Formatear el Excel
            from openpyxl.styles import PatternFill, Font, Alignment
            workbook = writer.book
            worksheet = writer.sheets['Reuniones']
            
            # Encabezado azul
            fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
            font = Font(bold=True, color='FFFFFF')
            
            for cell in worksheet[1]:
                cell.fill = fill
                cell.font = font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Ajustar anchos de columna
            worksheet.column_dimensions['A'].width = 10
            worksheet.column_dimensions['B'].width = 20
            worksheet.column_dimensions['C'].width = 25
        
        print(f"✅ Excel creado: {nombre_archivo}\n")
        
        if IN_COLAB:
            files.download(nombre_archivo)

if __name__ == "__main__":
    main()